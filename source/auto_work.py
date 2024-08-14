import re
import pandas as pd
from openpyxl import load_workbook
from collections import defaultdict

# 엑셀 파일 로드
workbook = load_workbook("C://Users//jjhee//OneDrive//바탕 화면//sins_work//comparison_results.xlsx", data_only=True)
sheet = workbook.active

# 정규 표현식
material_regex = re.compile(r'^성분:\s')
numbering_regex = re.compile(r'^\d{4}\s\d{4}-\d{2}-\d{2}')
area_regex = re.compile(r'^\d+')

material_name = None
data = {}
for row in sheet.iter_rows(min_row=3, min_col=1, max_col=4):
    cell_val = row[0].value
    if not cell_val : continue
    if material_regex.match(cell_val):
        material_name = cell_val[4:].strip()
        data[material_name] = pd.DataFrame(columns=['area'])
    elif numbering_regex.match(cell_val):
        numbering = cell_val[:4].strip()
        data[material_name].loc[numbering] = row[3].value

# 중복 제거된 조합 찾기
def find_combination(standard, unit_num):
    combination_set = set()
    for k, df in data.items():
        std_val = df.loc[standard]['area']
        diff_val = (df['area'] - std_val).abs()
        index = sorted(range(len(diff_val)), key=lambda x: diff_val[x], reverse=False)[1:unit_num+1]
        combination_set.add(tuple(sorted([df.index[x] for x in index])))
    
    return combination_set

# '2402', 0.05, 3
def calculate_error(standard, error, unit_num):
    combination_set = find_combination(standard, unit_num)

    output = {}
    for set in combination_set:
        temp_dict = {}
        for k, df in data.items():
            std_val = df.loc[standard]['area']
            avg = df.loc[[x for x in set]]['area'].mean()
            if (avg > std_val * (1+error)) or (avg < std_val * (1-error)) : break
            temp_dict[k] = avg/std_val - 1
        if len(data) == len(temp_dict) : output[','.join([x for x in set])] = temp_dict

    return output

# calculate_error('2402', 0.05, 3)