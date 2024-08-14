import re
import pandas as pd
from PyQt5.QtWidgets import (
    QApplication, QWidget, QLabel, QVBoxLayout, 
    QPushButton, QHBoxLayout, QTextBrowser, QLineEdit
)
from PyQt5.QtGui import QIcon
from openpyxl import load_workbook

route_first = "파일 : "

class MyApp(QWidget):

    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('신건호 외주')
        self.setGeometry(1000, 200, 500, 300)
        self.setAcceptDrops(True)

        # 초기 레이아웃 설정
        self.layout = QVBoxLayout(self)

        # 라벨 생성 및 레이아웃 추가
        self.label = QLabel(route_first, self)
        self.layout.addWidget(self.label)

        # 계산 버튼
        self.check_btn = QPushButton("확인", self)
        self.check_btn.clicked.connect(self.loadData)
        self.layout.addWidget(self.check_btn)

        self.setLayout(self.layout)

        self.show()

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.accept()
        else:
            event.ignore()

    def dropEvent(self, event):
        labelword = ""
        files = [u.toLocalFile() for u in event.mimeData().urls()]
        for f in files:
            print(f)
            labelword += route_first + f

        self.label.setText(labelword)

    def find_combination(self, standard, unit_num):
        combination_set = set()
        for k, df in self.data.items():
            std_val = df.loc[standard]['area']
            diff_val = (df['area'] - std_val).abs()
            index = sorted(range(len(diff_val)), key=lambda x: diff_val[x], reverse=False)[1:unit_num+1]
            combination_set.add(tuple(sorted([df.index[x] for x in index])))

        return combination_set

    def calculate_error(self, standard, error, unit_num):
        combination_set = self.find_combination(standard, unit_num)

        output = {}
        for set in combination_set:
            temp_dict = {}
            for k, df in self.data.items():
                std_val = df.loc[standard]['area']
                avg = df.loc[[x for x in set]]['area'].mean()
                if (avg > std_val * (1+error)) or (avg < std_val * (1-error)) : break
                temp_dict[k] = avg/std_val - 1
            if len(self.data) == len(temp_dict) : output[','.join([x for x in set])] = temp_dict

        str_list = []
        if output :
            # output 출력 텍스트 형식
            for num, v in output.items():
                str_list.append(num)
                for mat, cal in v.items():
                    str_list.append(f"{mat} : {cal}")
        else :
            str_list.append("없음")

        self.tb.append("\n".join(str_list))

    def loadData(self):
        path_regex = re.compile(r'^파일\s:\s[A-Za-z]:\/(?:[^\\/:*?"<>|\r\n]+\/)*[^\\/:*?"<>|\r\n]+\.xlsx$')
        path = self.label.text()
        if not path_regex.match(path):
            self.label.setText("파일이 잘못되었습니다.")
            return 

        # 엑셀 파일 로드
        workbook = load_workbook(path[5:].replace('/', '\\'), data_only=True)
        sheet = workbook.active

        # 정규 표현식
        material_regex = re.compile(r'^성분:\s')
        numbering_regex = re.compile(r'^\d{4}\s\d{4}-\d{2}-\d{2}')

        material_name = None
        data = {}
        for row in sheet.iter_rows(min_row=3, min_col=1, max_col=4):
            cell_val = row[0].value
            if not cell_val: continue
            if material_regex.match(cell_val):
                material_name = cell_val[4:].strip()
                data[material_name] = pd.DataFrame(columns=['area'])
            elif numbering_regex.match(cell_val):
                numbering = cell_val[:4].strip()
                data[material_name].loc[numbering] = row[3].value

        self.data = data
        self.analyse_page()

    def analyse_page(self):
        self.label.hide()
        self.check_btn.hide()
        self.resize(800, 500)

        self.numbering_list = []
        x_point = 100
        y_point = 100

        for item in list(self.data.values())[0].index:
            button = QPushButton(str(item), self)
            button.move(x_point, y_point)
            button.clicked.connect(lambda _, b=item: self.calculate_error(b, 0.05, 3))
            y_point += 30
            button.show()
            self.numbering_list.append(button)

        self.tb = QTextBrowser(self)
        self.tb.setAcceptRichText(True)
        self.tb.setOpenExternalLinks(True)
        self.tb.move(300, 180)
        self.tb.setFixedSize(500, 300)
        self.tb.show()

if __name__ == '__main__':
    app = QApplication([])
    myapp = MyApp()
    app.exec_()